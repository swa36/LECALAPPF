import os
from datetime import datetime
from typing import Dict

from catalog.models import (
    Category,
    NameAdditionalAttributes,
    Prices,
    Product,
    TypePrices,
    ValueAdditionalAttributes,
)
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from src.lekala_class.class_1C.ExChange1C import ExChange1C


class GetData1C(ExChange1C):
    def __init__(self):
        super().__init__()

    def set_name_attribute(self):
        data_name_attrib = self.get_name_additional_attributes()["value"]
        for item in data_name_attrib:
            NameAdditionalAttributes.objects.update_or_create(
                uuid_1C=item["ref_key"],
                defaults={"name_attribute": item["description"]},
            )

    def set_category_catalog(self):
        data_category = self.get_category()["value"]
        dict_cat = {
            "main_cat": [],
            "sub_cat": [],
            "sub_sub_cat": [],
            "sub_sub_sub_cat": [],
        }

        for item in data_category:
            if item["parent_key"] == "00000000-0000-0000-0000-000000000000":
                dict_cat["main_cat"].append(item)

        list_id_main_cat = [item["ref_key"] for item in dict_cat["main_cat"]]

        for item in data_category:
            if item["parent_key"] in list_id_main_cat:
                dict_cat["sub_cat"].append(item)

        list_id_sub_cat = [item["ref_key"] for item in dict_cat["sub_cat"]]

        for item in data_category:
            if item["parent_key"] in list_id_sub_cat:
                dict_cat["sub_sub_cat"].append(item)

        list_id_sub_sub_cat = [item["ref_key"] for item in dict_cat["sub_sub_cat"]]

        for item in data_category:
            if item["parent_key"] in list_id_sub_sub_cat:
                dict_cat["sub_sub_sub_cat"].append(item)

        for group, values in dict_cat.items():
            for item in values:
                defaults = {"name": item["description"]}
                if group != "main_cat":
                    defaults["parent"] = Category.objects.get(uuid_1C=item["parent_key"])
                Category.objects.update_or_create(
                    uuid_1C=item["ref_key"],
                    defaults=defaults,
                )

    def set_type_price(self):
        suffix_type_price = {
            "Закупочная": "cost_price",
            "Оптовая": "wholesale_price",
            "Опт2": "wholesale_price_2",
            "Опт3": "wholesale_price_3",
            "Розничная": "retail_price",
        }
        data_type_price = self.get_type_price()["value"]
        for item in data_type_price:
            TypePrices.objects.update_or_create(
                uuid_1C=item["ref_key"],
                defaults={
                    "type_price": item["description"],
                    "suffix": suffix_type_price[item["description"]],
                },
            )

    def save_image_errors_to_excel(self, data_list, file_path="ошибки_изображений.xlsx"):
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Ошибки загрузки"
            headers = ["Наименование товара", "Артикул в 1С"]
            ws.append(headers)

        for row in data_list:
            ws.append(row[:2])

        for col in ws.iter_cols(min_row=1, max_row=ws.max_row, max_col=2):
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_length + 2

        wb.save(file_path)

    def get_quantity_in_order(self) -> Dict[str, int]:
        quantities: Dict[str, int] = {}
        try:
            orders = self.get_orders_in_shipping().get("value", [])
        except Exception as exc:
            print(f"Ошибка при получении резервов: {exc}")
            return quantities

        for order in orders:
            for item in order.get("items", []):
                try:
                    product_key = str(item["nomenclature_key"])
                    quantity = int(item["quantity"])
                except (KeyError, ValueError, TypeError) as exc:
                    print(f"Ошибка количества в резерве: {exc}")
                    continue
                quantities[product_key] = quantities.get(product_key, 0) + quantity

        return quantities

    def set_catalog_data_stock(self, data_catalog):
        type_prices = TypePrices.objects.all()

        for item in data_catalog:
            stock = max(int(item.get("in_stock") or 0), 0)
            product_before = Product.objects.filter(uuid_1C=item["ref_key"]).first()
            should_refresh_details = (
                product_before is None
                or product_before.data_version != item.get("data_version")
            )

            product, created = Product.objects.update_or_create(
                uuid_1C=item["ref_key"],
                defaults={
                    "article_1C": item["article"].strip(),
                    "code_1C": item["code"],
                    "data_version": item.get("data_version") or "",
                    "name": item["description"].strip(),
                    "description": item["description_text"],
                    "stock": stock,
                    "main_img_uuid": item.get("picture_file_key") or None,
                    "category": Category.objects.get(uuid_1C=item["parent_key"]),
                },
            )

            price_by_type = {
                str(price["price_type_key"]): price.get("price", 0)
                for price in item.get("prices", [])
            }
            price_dict = {
                type_price.suffix: price_by_type.get(str(type_price.uuid_1C), 0)
                for type_price in type_prices
            }
            Prices.objects.update_or_create(product=product, defaults=price_dict)

            if should_refresh_details:
                print(
                    f"Создана/обновлена номенклатура "
                    f"{product.name} {product.article_1C}"
                )
                if created:
                    os.makedirs("logs", exist_ok=True)
                    with open("logs/new_item.log", "a", encoding="utf-8") as log_file:
                        log_file.write(
                            f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\t"
                            f"{product.name}\t{product.article_1C}\t{product.code_1C}\n"
                        )

                for attribute in item.get("additional_attributes", []):
                    ValueAdditionalAttributes.objects.update_or_create(
                        product=product,
                        attribute_name=NameAdditionalAttributes.objects.get(
                            uuid_1C=attribute["property_key"]
                        ),
                        defaults={"value_attribute": attribute["value"]},
                    )

                self.get_img(product.uuid_1C)
